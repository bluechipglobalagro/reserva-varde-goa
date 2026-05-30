from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FOREST = "1F3A2E"
GOLD = "B08A3E"
IVORY = "FBF8F1"
BEIGE = "F4EDDD"
CHAR = "2A2A2A"
LINE = "C7B97A"

thin = Side(border_style="thin", color="C9C2A8")
bord = Border(top=thin, left=thin, right=thin, bottom=thin)

def header_fmt(cell):
    cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=FOREST)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = bord

def body_fmt(cell, alt=False, bold=False, total=False):
    if total:
        cell.fill = PatternFill("solid", fgColor=BEIGE)
        cell.font = Font(name="Calibri", size=11, bold=True, color=FOREST)
    else:
        cell.fill = PatternFill("solid", fgColor=IVORY if alt else "FFFFFF")
        cell.font = Font(name="Calibri", size=11, bold=bold, color=CHAR)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = bord

def title_fmt(cell):
    cell.font = Font(name="Calibri", size=16, bold=True, color=FOREST)
    cell.alignment = Alignment(horizontal="left", vertical="center")

def subtitle_fmt(cell):
    cell.font = Font(name="Calibri", size=11, italic=True, color=GOLD)
    cell.alignment = Alignment(horizontal="left", vertical="center")

def write_sheet(ws, title, subtitle, schedule, totalsRowIdx=None):
    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 48
    ws.merge_cells('A1:D1')
    title_fmt(ws['A1'])
    ws['A1'] = title
    ws.merge_cells('A2:D2')
    subtitle_fmt(ws['A2'])
    ws['A2'] = subtitle
    start = 4
    for i, row in enumerate(schedule):
        r = start + i
        is_header = i == 0
        is_total = (totalsRowIdx is not None and i == totalsRowIdx) or row[0].upper().startswith("TOTAL")
        for j, val in enumerate(row):
            c = ws.cell(row=r, column=j+1, value=val)
            if is_header:
                header_fmt(c)
            else:
                body_fmt(c, alt=(i % 2 == 0), total=is_total)
        ws.row_dimensions[r].height = 22 if not is_header else 28
    ws.freeze_panes = ws['A5']

wb = Workbook()
ws = wb.active
ws.title = "2BHK 2500 sq.ft"

schedule_2 = [
  ["Room / Space", "Area (sq.ft)", "Function", "Notes"],
  ["Living Room", "320", "Public lounge", "40 ft module, double-height ceiling, plantation view"],
  ["Dining Area", "180", "Family dining + entertaining", "Open to living, connector zone"],
  ["Kitchen", "180", "Cooking + prep", "20 ft module with island, low-flow fixtures"],
  ["Family Lounge / Study", "160", "Reading, WFH, second TV", "Convertible space"],
  ["Master Bedroom", "260", "Primary suite", "40 ft module, garden-facing window wall"],
  ["Master Toilet", "70", "Attached bath", "Walk-in shower, skylight"],
  ["Bedroom 2", "220", "Guest / family", "40 ft module, courtyard view"],
  ["Bedroom 2 Toilet", "60", "Attached bath", "Compact wet zone"],
  ["Powder Toilet", "35", "Guest WC", "Off entry foyer"],
  ["Utility / Laundry", "70", "Service", "Greywater point, washing"],
  ["Storage", "55", "Pantry + general", "Walk-in, ventilated"],
  ["Internal Courtyard", "260", "Open-to-sky core", "Microclimate cooling, water feature"],
  ["Covered Deck / Verandah", "320", "Outdoor living", "Pergola roof, plantation-facing"],
  ["Outdoor Sit-Out", "180", "Garden lounge", "Stone deck, optional plunge pool"],
  ["Corridor / Connector", "100", "Glass link bridges", "Shaded transitional spine"],
  ["Service Zone", "40", "Bins, gas, panel", "Concealed behind louvres"],
  ["TOTAL", "2,505", "—", "Approximate (±2% architectural tolerance)"],
]
write_sheet(ws, "2BHK Compact Luxury Modular Eco Estate — Room Schedule",
            "Reserva Verde Goa  ·  Built-up 2,500 Sq.Ft  ·  1-Acre Plantation Estate",
            schedule_2)

ws2 = wb.create_sheet("3BHK 3500 sq.ft")
schedule_3 = [
  ["Room / Space", "Area (sq.ft)", "Function", "Notes"],
  ["Living Room", "380", "Grand public lounge", "Two 40 ft modules linked, double-height central bay"],
  ["Dining Area", "220", "8–10 seater", "Plantation-facing glass wall"],
  ["Large Kitchen", "220", "Open island + prep", "40 ft module, breakfast counter"],
  ["Family Lounge", "200", "Casual second living", "Sliding doors to courtyard"],
  ["Study / WFH Room", "160", "Work-from-estate", "Acoustic insulation, garden view"],
  ["Master Bedroom", "300", "Primary suite", "40 ft module, walk-in wardrobe alcove"],
  ["Master Toilet", "90", "Spa-style", "Twin vanity, rain shower"],
  ["Bedroom 2", "240", "Guest / kids", "40 ft module"],
  ["Bedroom 2 Toilet", "70", "Attached", "Compact wet zone"],
  ["Bedroom 3", "240", "Guest / family", "40 ft module"],
  ["Bedroom 3 Toilet", "70", "Attached", "Compact wet zone"],
  ["Powder Toilet", "40", "Guest WC", "Off foyer"],
  ["Utility / Laundry", "90", "Service", "Greywater + washing"],
  ["Storage / Pantry", "80", "Walk-in", "Ventilated, mosquito-screened"],
  ["Internal Courtyard", "340", "Open-to-sky", "Stone water channel, native planting"],
  ["Covered Verandah", "300", "Shaded outdoor living", "Pergola + bamboo screen"],
  ["Outdoor Deck", "220", "Wellness / dining deck", "Timber deck on stilts"],
  ["Plantation-Facing Sit-Out", "150", "Quiet corner", "Hammock zone"],
  ["Corridor / Connector", "140", "Glass-link spine", "Connects wings"],
  ["Service Zone", "50", "Bins, gas, panel", "Side service yard"],
  ["TOTAL", "3,500", "—", "Approximate (±2% architectural tolerance)"],
]
write_sheet(ws2, "3BHK Premium Modular Agro Estate — Room Schedule",
            "Reserva Verde Goa  ·  Built-up 3,500 Sq.Ft  ·  1-Acre Plantation Estate",
            schedule_3)

ws3 = wb.create_sheet("4BHK 4500 sq.ft")
schedule_4 = [
  ["Room / Space", "Area (sq.ft)", "Function", "Notes"],
  ["Grand Living Room", "460", "Pavilion-style lounge", "Two 40 ft modules + glass extension, plantation view"],
  ["Formal Dining", "260", "10–12 seater", "Adjacent to dirty kitchen"],
  ["Open Kitchen", "220", "Show kitchen + island", "40 ft module"],
  ["Dirty / Service Kitchen", "140", "Heavy cooking + storage", "20 ft module, exhaust ventilation"],
  ["Family Lounge", "240", "Casual TV / second living", "Connects to wellness deck"],
  ["Study / Office", "180", "WFH / library", "Sound-insulated"],
  ["Master Suite Bedroom", "340", "King suite + reading nook", "40 ft module + extension"],
  ["Master Walk-in Wardrobe", "120", "Dressing", "Ventilated"],
  ["Master Toilet", "120", "Spa bath", "Twin vanity, rain shower, soaking tub"],
  ["Bedroom 2", "260", "Guest suite", "40 ft module"],
  ["Bedroom 2 Toilet", "80", "Attached", "Walk-in shower"],
  ["Bedroom 3", "260", "Family bedroom", "40 ft module"],
  ["Bedroom 3 Toilet", "80", "Attached", "Walk-in shower"],
  ["Bedroom 4", "240", "Family / kids", "40 ft module"],
  ["Bedroom 4 Toilet", "70", "Attached", "Compact wet zone"],
  ["Powder Toilet", "45", "Guest WC", "Off arrival foyer"],
  ["Utility / Laundry", "120", "Service", "Greywater + drying yard"],
  ["Storage / Pantry", "100", "Walk-in", "Climate-controlled"],
  ["Staff / Caretaker Room (opt.)", "140", "Live-in caretaker", "Separate entry, attached toilet"],
  ["Central Courtyard", "420", "Open-to-sky core", "Reflective pool + tree"],
  ["Covered Deck", "320", "Entertainment deck", "Pergola + outdoor kitchen counter"],
  ["Yoga / Wellness Deck", "200", "Meditation + yoga", "Bamboo screen, sunrise-facing"],
  ["Outdoor Dining Pavilion", "180", "Al fresco dining", "Pergola roof"],
  ["Private Garden Court", "200", "Master-suite private garden", "Walled, fragrant planting"],
  ["Corridor / Connector", "170", "Glass-link spine", "Connects 3 wings"],
  ["Service Zone", "60", "Bins, gas, panel", "Concealed service yard"],
  ["TOTAL", "4,525", "—", "Approximate (±2% architectural tolerance)"],
]
write_sheet(ws3, "4BHK Signature Modular Luxury Estate — Room Schedule",
            "Reserva Verde Goa  ·  Built-up 4,500 Sq.Ft  ·  1-Acre Plantation Estate",
            schedule_4)

ws4 = wb.create_sheet("Specs Matrix")
matrix = [
  ["Attribute", "2BHK", "3BHK", "4BHK", "Customisable"],
  ["Asset type", "GL Modular Container Eco Home + Plantation", "GL Modular Container Eco Home + Plantation", "GL Modular Container Eco Home + Plantation", "GL Modular Container Eco Home + Plantation"],
  ["Built-up area", "2,500 sq.ft", "3,500 sq.ft", "4,500 sq.ft", "Customisable (1,500–6,000)"],
  ["Estate size", "1 acre", "1 acre", "1 acre", "1 acre"],
  ["Storeys", "G only", "G only", "G only", "G only"],
  ["Bedrooms", "2 + powder", "3 + powder", "4 + powder + opt. staff", "1–4 + powder"],
  ["Bathrooms", "2 + powder", "3 + powder", "4 + powder", "Matched to BR count"],
  ["40 ft modules", "4", "6", "8", "1–12"],
  ["20 ft modules", "2", "2", "3", "0–6"],
  ["Layout intent", "Courtyard", "L-shape pavilion", "Three-wing pavilion", "Buyer-selected"],
  ["Recommended placement", "Front-to-centre", "Centre", "Rear-of-centre", "Any"],
  ["Plantation share of plot", "~63%", "~57%", "~50%", "≥45%"],
  ["RWH capacity", "~15,000 L", "~25,000 L", "~40,000 L", "Sized to design"],
  ["Solar-ready kWp", "5–7", "8–10", "12–15", "Sized to design"],
  ["Greywater reuse", "Yes", "Yes (reed bed)", "Yes (wetland)", "Yes"],
  ["Bio-septic", "Single chamber", "Twin chamber", "Twin + filter pond", "Sized to design"],
  ["Pool option", "Plunge", "Plunge / lap", "Swimming pool", "Buyer-selected"],
  ["Rental readiness", "Optional", "Optional", "Optional", "Off / Light / Full"],
  ["Ideal buyer", "Couples, NRIs, eco buyers", "HNI families, NRIs, wellness", "HNIs, retreat ops, hospitality", "Any persona"],
]
ws4.column_dimensions['A'].width = 28
for col in 'BCDE':
    ws4.column_dimensions[col].width = 32
ws4.merge_cells('A1:E1')
title_fmt(ws4['A1'])
ws4['A1'] = "Reserva Verde Goa — Final Specifications Matrix"
ws4.merge_cells('A2:E2')
subtitle_fmt(ws4['A2'])
ws4['A2'] = "All values concept-stage. Subject to final architectural planning & eco-sensitive development guidelines."

for i, row in enumerate(matrix):
    r = 4 + i
    is_header = i == 0
    for j, val in enumerate(row):
        c = ws4.cell(row=r, column=j+1, value=val)
        if is_header: header_fmt(c)
        else: body_fmt(c, alt=(i % 2 == 0))
    ws4.row_dimensions[r].height = 24 if not is_header else 28
ws4.freeze_panes = ws4['B5']

ws5 = wb.create_sheet("Site Zoning")
zoning_data = [
    ["Zone", "2BHK (sq.ft)", "2BHK %", "3BHK (sq.ft)", "3BHK %", "4BHK (sq.ft)", "4BHK %"],
    ["Entry & Arrival", "2,200", "5.0%", "2,600", "6.0%", "3,200", "7.3%"],
    ["Container Home (built-up)", "2,500", "5.7%", "3,500", "8.0%", "4,500", "10.3%"],
    ["Outdoor Lifestyle", "3,800", "8.7%", "4,600", "10.5%", "5,800", "13.3%"],
    ["Plantation Zone", "27,500", "63.0%", "25,000", "57.4%", "22,000", "50.5%"],
    ["Water & Sustainability", "1,800", "4.1%", "2,000", "4.6%", "2,200", "5.1%"],
    ["Utility Zone", "900", "2.1%", "1,200", "2.8%", "1,500", "3.4%"],
    ["Landscape Buffer + Path", "4,860", "11.2%", "4,660", "10.7%", "4,360", "10.0%"],
    ["TOTAL", "43,560", "100%", "43,560", "100%", "43,560", "100%"],
]
ws5.column_dimensions['A'].width = 30
for col in 'BCDEFG':
    ws5.column_dimensions[col].width = 14
ws5.merge_cells('A1:G1')
title_fmt(ws5['A1'])
ws5['A1'] = "1-Acre Site Zoning Comparison"
ws5.merge_cells('A2:G2')
subtitle_fmt(ws5['A2'])
ws5['A2'] = "Reserva Verde Goa  ·  43,560 sq.ft per estate  ·  Zone allocation by model"

for i, row in enumerate(zoning_data):
    r = 4 + i
    is_header = i == 0
    is_total = row[0].upper() == "TOTAL"
    for j, val in enumerate(row):
        c = ws5.cell(row=r, column=j+1, value=val)
        if is_header: header_fmt(c)
        else: body_fmt(c, alt=(i % 2 == 0), total=is_total)
    ws5.row_dimensions[r].height = 24 if not is_header else 28

ws6 = wb.create_sheet("Container Modules")
mods = [
    ["Model", "40 ft HC modules", "20 ft modules", "Connector / decks (sq.ft)", "Approx. built-up (sq.ft)"],
    ["2BHK / 2,500", "4", "2", "Glass spine + verandah + deck (~900)", "2,500"],
    ["3BHK / 3,500", "6", "2", "L-spine + verandah + deck (~1,260)", "3,500"],
    ["4BHK / 4,500", "8", "3", "Spine + pavilion deck + court roof (~1,485)", "4,500"],
    ["Customisable", "1–12", "0–6", "Tailored", "1,500 – 6,000"],
]
ws6.column_dimensions['A'].width = 24
for col, w in zip('BCDE', [18, 18, 42, 30]):
    ws6.column_dimensions[col].width = w
ws6.merge_cells('A1:E1')
title_fmt(ws6['A1'])
ws6['A1'] = "Container Module Strategy"
ws6.merge_cells('A2:E2')
subtitle_fmt(ws6['A2'])
ws6['A2'] = "40 ft HC (40×8 / 320 sq.ft) primary · 20 ft for service & wet zones"
for i, row in enumerate(mods):
    r = 4 + i
    is_header = i == 0
    for j, val in enumerate(row):
        c = ws6.cell(row=r, column=j+1, value=val)
        if is_header: header_fmt(c)
        else: body_fmt(c, alt=(i % 2 == 0))
    ws6.row_dimensions[r].height = 24

wb.save("/sessions/relaxed-quirky-pasteur/mnt/outputs/Reserva_Verde_Goa_Schedules.xlsx")
print("Saved Reserva_Verde_Goa_Schedules.xlsx")
